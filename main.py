"""Organizador y mantenimiento automático de archivos para la PC."""
import hashlib
import json
import logging
import os
import shutil
import sys
import tempfile
import time
from pathlib import Path

CONFIG_PATH = Path(__file__).parent / "config.json"
LOG_PATH = Path(__file__).parent / "organizador.log"
CATEGORIA_OTROS = "Otros"
CATEGORIA_DOCUMENTOS = "Documentos"
CATEGORIA_SIN_CLASIFICAR = "Sin_Clasificar"
MODELO_CLAUDE = "claude-haiku-4-5"
MAX_PAGINAS_PDF = 3
MAX_PARRAFOS_DOCX = 50
MAX_FILAS_EXCEL = 30
MAX_HOJAS_EXCEL = 3
MAX_CARACTERES_TEXTO = 3000
MAX_LARGO_NOMBRE_CATEGORIA = 40
CARACTERES_INVALIDOS_CARPETA = '<>:"/\\|?*'
EXTENSIONES_DESCARGA_EN_CURSO = {"crdownload", "tmp", "part", "partial", "download"}
EXTENSIONES_IGNORADAS = {"lnk", "url", "ini"}
SEGUNDOS_MINIMOS_DESDE_MODIFICACION = 3

CARPETA_REVISION_DUPLICADOS = Path.home() / "Desktop" / "revision_duplicados"

def _carpetas_temporales() -> list[Path]:
    candidatas = [
        Path(os.environ.get("TEMP", tempfile.gettempdir())),
        Path(os.environ.get("TMP", tempfile.gettempdir())),
        (Path(os.environ["LOCALAPPDATA"]) / "Temp") if os.environ.get("LOCALAPPDATA") else None,
    ]
    vistas = set()
    resultado = []
    for carpeta in candidatas:
        if not carpeta or not carpeta.exists():
            continue
        resuelta = carpeta.resolve()
        if resuelta not in vistas:
            vistas.add(resuelta)
            resultado.append(carpeta)
    return resultado


CARPETAS_TEMPORALES = _carpetas_temporales()


def configurar_logging() -> None:
    if logging.getLogger().handlers:
        return  # ya configurado (evita duplicar handlers si se llama más de una vez)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(LOG_PATH, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def cargar_reglas(config_path: Path) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def obtener_carpeta_descargas() -> Path:
    return Path(os.path.expanduser("~")) / "Downloads"


def obtener_carpetas_a_organizar(reglas: dict) -> list[Path]:
    home = Path(os.path.expanduser("~"))
    return [home / nombre for nombre in reglas.get("carpetas_a_organizar", ["Downloads"])]


def construir_mapa_extensiones(reglas: dict[str, list[str]]) -> dict[str, str]:
    mapa = {}
    for categoria, extensiones in reglas.items():
        for ext in extensiones:
            mapa[ext.lower()] = categoria
    return mapa


def categoria_para_archivo(archivo: Path, mapa_extensiones: dict[str, str]) -> str:
    ext = archivo.suffix.lower().lstrip(".")
    return mapa_extensiones.get(ext, CATEGORIA_OTROS)


def destino_sin_colision(carpeta_destino: Path, nombre_archivo: str) -> Path:
    destino = carpeta_destino / nombre_archivo
    if not destino.exists():
        return destino

    base = Path(nombre_archivo).stem
    ext = Path(nombre_archivo).suffix
    contador = 1
    while True:
        candidato = carpeta_destino / f"{base} ({contador}){ext}"
        if not candidato.exists():
            return candidato
        contador += 1


def extraer_texto_pdf(archivo: Path) -> str | None:
    try:
        from pypdf import PdfReader

        lector = PdfReader(str(archivo))
        paginas = lector.pages[:MAX_PAGINAS_PDF]
        texto = "\n".join(pagina.extract_text() or "" for pagina in paginas)
        return texto.strip() or None
    except Exception:
        return None


def extraer_texto_docx(archivo: Path) -> str | None:
    try:
        from docx import Document

        documento = Document(str(archivo))
        parrafos = [
            p.text for p in documento.paragraphs[:MAX_PARRAFOS_DOCX] if p.text.strip()
        ]
        return "\n".join(parrafos).strip() or None
    except Exception:
        return None


def extraer_texto_xlsx(archivo: Path) -> str | None:
    libro = None
    try:
        from openpyxl import load_workbook

        libro = load_workbook(str(archivo), read_only=True, data_only=True)
        partes = []
        for hoja in libro.worksheets[:MAX_HOJAS_EXCEL]:
            partes.append(f"[Hoja: {hoja.title}]")
            for fila in hoja.iter_rows(max_row=MAX_FILAS_EXCEL, values_only=True):
                valores = [str(v) for v in fila if v is not None and str(v).strip()]
                if valores:
                    partes.append(" | ".join(valores))
        return "\n".join(partes).strip() or None
    except Exception:
        return None
    finally:
        if libro is not None:
            libro.close()


def extraer_texto_xls(archivo: Path) -> str | None:
    try:
        import xlrd

        libro = xlrd.open_workbook(str(archivo))
        partes = []
        for hoja in libro.sheets()[:MAX_HOJAS_EXCEL]:
            partes.append(f"[Hoja: {hoja.name}]")
            for fila_idx in range(min(hoja.nrows, MAX_FILAS_EXCEL)):
                valores = [
                    str(v) for v in hoja.row_values(fila_idx) if str(v).strip()
                ]
                if valores:
                    partes.append(" | ".join(valores))
        return "\n".join(partes).strip() or None
    except Exception:
        return None


def extraer_texto(archivo: Path) -> str | None:
    ext = archivo.suffix.lower().lstrip(".")
    if ext == "pdf":
        return extraer_texto_pdf(archivo)
    if ext == "docx":
        return extraer_texto_docx(archivo)
    if ext == "xlsx":
        return extraer_texto_xlsx(archivo)
    if ext == "xls":
        return extraer_texto_xls(archivo)
    return None


def listar_categorias_existentes(carpeta_documentos: Path) -> list[str]:
    if not carpeta_documentos.is_dir():
        return []
    return sorted(
        p.name
        for p in carpeta_documentos.iterdir()
        if p.is_dir() and p.name != CATEGORIA_SIN_CLASIFICAR
    )


def sanitizar_nombre_categoria(nombre: str) -> str:
    limpio = nombre.strip()
    for caracter in CARACTERES_INVALIDOS_CARPETA:
        limpio = limpio.replace(caracter, "_")
    limpio = limpio.strip(" .")[:MAX_LARGO_NOMBRE_CATEGORIA].strip()
    return limpio


def clasificar_categoria(
    nombre_archivo: str, texto: str | None, categorias_existentes: list[str]
) -> str | None:
    try:
        import anthropic

        cliente = anthropic.Anthropic()
        categorias_texto = ", ".join(categorias_existentes) if categorias_existentes else "(ninguna todavía)"

        if texto:
            seccion_contenido = f"Contenido (primeras páginas):\n{texto[:MAX_CARACTERES_TEXTO]}"
        else:
            seccion_contenido = (
                "No se pudo extraer el contenido de este archivo (por ejemplo, es un "
                ".pptx u otro formato no soportado para lectura, o la extracción "
                "falló). Guiate ÚNICAMENTE por el nombre del archivo para inferir el "
                "tema — muchas veces el nombre ya lo indica (ej: 'Clase 6 - Algebra.pptx' "
                "es de la materia Algebra)."
            )

        prompt = (
            f"Nombre del archivo: {nombre_archivo}\n\n"
            f"{seccion_contenido}\n\n"
            f"Carpetas que ya existen dentro de Documentos: {categorias_texto}\n\n"
            "Estás organizando la carpeta Documentos de un usuario que mezcla archivos "
            "académicos (de la facultad) y personales (trámites, certificados, CV, etc). "
            "Elegí la carpeta que mejor represente el TEMA o TIPO de este documento. "
            "Puede ser una materia universitaria (ej: 'Calculo I', 'Bases de Datos'), "
            "un tipo de documento personal (ej: 'CV', 'CUIL', 'Certificados', "
            "'Constancias'), o cualquier otra categoría clara y específica que "
            "corresponda.\n\n"
            "Si el documento corresponde EXACTAMENTE al mismo tema o tipo de "
            "documento que una de las carpetas que ya existen, respondé ese mismo "
            "nombre (para no crear una carpeta duplicada). No reutilices una carpeta "
            "solo porque el tema es parecido: una constancia de CUIL y una constancia "
            "de alumno, por ejemplo, son tipos de documento distintos aunque ambas "
            "sean 'constancias' — no las mezcles en la misma carpeta. Si no hay una "
            "carpeta que coincida exactamente pero podés identificar "
            "claramente de qué se trata (por contenido o por el nombre del archivo), "
            "inventá un nombre corto y específico (2 a 4 palabras, en español, sin "
            "caracteres especiales, apto para nombre de carpeta, máximo "
            f"{MAX_LARGO_NOMBRE_CATEGORIA} caracteres). Si es genérico o ambiguo pero "
            "reconocible respondé exactamente "
            f"'{CATEGORIA_OTROS}'. Si ni el contenido ni el nombre del archivo dan "
            "ninguna pista de qué se trata, respondé exactamente "
            f"'{CATEGORIA_SIN_CLASIFICAR}'.\n\n"
            "Respondé ÚNICAMENTE con el nombre de la carpeta: una etiqueta corta, "
            "nunca una oración ni una explicación."
        )
        respuesta = cliente.messages.create(
            model=MODELO_CLAUDE,
            max_tokens=20,
            messages=[{"role": "user", "content": prompt}],
        )
        texto_respuesta = next(
            (b.text for b in respuesta.content if b.type == "text"), ""
        ).strip()
    except Exception as e:
        logging.warning(f"Fallo la clasificación por IA de '{nombre_archivo}': {e}")
        return None

    for categoria in categorias_existentes:
        if categoria.lower() == texto_respuesta.lower():
            return categoria

    return sanitizar_nombre_categoria(texto_respuesta) or None


def clasificar_documento(archivo: Path, carpeta_documentos: Path) -> str:
    texto = extraer_texto(archivo)
    categorias_existentes = listar_categorias_existentes(carpeta_documentos)
    categoria = clasificar_categoria(archivo.name, texto, categorias_existentes)
    return categoria or CATEGORIA_SIN_CLASIFICAR


def archivo_listo_para_mover(archivo: Path) -> bool:
    ext = archivo.suffix.lower().lstrip(".")
    if ext in EXTENSIONES_DESCARGA_EN_CURSO:
        return False

    try:
        segundos_desde_modificacion = time.time() - archivo.stat().st_mtime
    except FileNotFoundError:
        return False

    return segundos_desde_modificacion >= SEGUNDOS_MINIMOS_DESDE_MODIFICACION


def organizar_carpeta(carpeta: Path, reglas: dict) -> None:
    if not carpeta.is_dir():
        return

    mapa_extensiones = construir_mapa_extensiones(reglas["categorias"])

    archivos = [
        f
        for f in carpeta.iterdir()
        if f.is_file()
        and f.suffix.lower().lstrip(".") not in EXTENSIONES_IGNORADAS
        and archivo_listo_para_mover(f)
    ]

    for archivo in archivos:
        categoria = categoria_para_archivo(archivo, mapa_extensiones)

        if categoria == CATEGORIA_DOCUMENTOS:
            carpeta_documentos = carpeta / CATEGORIA_DOCUMENTOS
            subcarpeta = clasificar_documento(archivo, carpeta_documentos)
            carpeta_destino = carpeta_documentos / subcarpeta
        else:
            carpeta_destino = carpeta / categoria

        carpeta_destino.mkdir(parents=True, exist_ok=True)
        destino = destino_sin_colision(carpeta_destino, archivo.name)
        shutil.move(str(archivo), str(destino))
        logging.info(f"Movido: {archivo.name} -> {carpeta_destino.relative_to(carpeta)}/{destino.name}")


# ----------------------------------------------------------------------
# MANTENIMIENTO: temporales, duplicados, programas instalados
# ----------------------------------------------------------------------


def tamano_legible(cantidad_bytes: float) -> str:
    for unidad in ["B", "KB", "MB", "GB"]:
        if cantidad_bytes < 1024:
            return f"{cantidad_bytes:.1f} {unidad}"
        cantidad_bytes /= 1024
    return f"{cantidad_bytes:.1f} TB"


def limpiar_temporales(ejecutar: bool) -> None:
    total_liberado = 0
    total_elementos = 0

    for carpeta in CARPETAS_TEMPORALES:
        logging.info(f"Revisando temporales en: {carpeta}")
        for item in carpeta.glob("*"):
            try:
                if item.is_file():
                    tamano = item.stat().st_size
                    if ejecutar:
                        item.unlink()
                    total_liberado += tamano
                    total_elementos += 1
                elif item.is_dir():
                    tamano = sum(f.stat().st_size for f in item.rglob("*") if f.is_file())
                    if ejecutar:
                        shutil.rmtree(item, ignore_errors=True)
                    total_liberado += tamano
                    total_elementos += 1
            except (PermissionError, OSError):
                continue  # archivo en uso, se saltea sin cortar el script

    accion = "Liberados" if ejecutar else "Se liberarían (simulación)"
    logging.info(f"Temporales: {accion} {tamano_legible(total_liberado)} en {total_elementos} elementos")


def hash_archivo(archivo: Path, bloque: int = 65536) -> str | None:
    h = hashlib.md5()
    try:
        with open(archivo, "rb") as f:
            while chunk := f.read(bloque):
                h.update(chunk)
        return h.hexdigest()
    except (PermissionError, OSError):
        return None


def buscar_duplicados(ejecutar: bool, carpetas: list[Path]) -> None:
    hashes: dict[str, Path] = {}
    duplicados: list[tuple[Path, Path]] = []

    for carpeta in carpetas:
        if not carpeta.is_dir():
            continue
        logging.info(f"Escaneando duplicados en: {carpeta}")
        for archivo in carpeta.rglob("*"):
            if not archivo.is_file() or CARPETA_REVISION_DUPLICADOS in archivo.parents:
                continue
            h = hash_archivo(archivo)
            if h is None:
                continue
            if h in hashes:
                duplicados.append((archivo, hashes[h]))
            else:
                hashes[h] = archivo

    logging.info(f"Duplicados: encontrados {len(duplicados)} archivos")

    if not duplicados:
        return

    if not ejecutar:
        for dup, original in duplicados:
            logging.info(f"  {dup} -> duplicado de {original}")
        return

    CARPETA_REVISION_DUPLICADOS.mkdir(parents=True, exist_ok=True)
    for dup, original in duplicados:
        try:
            destino = destino_sin_colision(CARPETA_REVISION_DUPLICADOS, dup.name)
            shutil.move(str(dup), str(destino))
            logging.info(f"Duplicado movido a revisión: {dup.name} (copia de {original.name})")
        except (PermissionError, OSError) as e:
            logging.warning(f"No se pudo mover duplicado {dup.name}: {e}")
    logging.info(f"Revisá y borrá manualmente lo que no necesites en: {CARPETA_REVISION_DUPLICADOS}")


def listar_programas() -> None:
    if sys.platform != "win32":
        logging.info("El listado de programas instalados solo funciona en Windows.")
        return

    import winreg

    rutas = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
    ]

    programas = []
    for hive, ruta in rutas:
        try:
            with winreg.OpenKey(hive, ruta) as clave:
                for i in range(winreg.QueryInfoKey(clave)[0]):
                    try:
                        subclave_nombre = winreg.EnumKey(clave, i)
                        with winreg.OpenKey(clave, subclave_nombre) as subclave:
                            nombre = winreg.QueryValueEx(subclave, "DisplayName")[0]
                            try:
                                fecha = winreg.QueryValueEx(subclave, "InstallDate")[0]
                            except FileNotFoundError:
                                fecha = "?"
                            programas.append((nombre, fecha))
                    except (FileNotFoundError, OSError):
                        continue
        except FileNotFoundError:
            continue

    programas = sorted(set(programas))
    for nombre, fecha in programas:
        logging.info(f"  {nombre}  (instalado: {fecha})")
    logging.info(f"Programas instalados: {len(programas)} en total")


def main() -> None:
    configurar_logging()
    reglas = cargar_reglas(CONFIG_PATH)

    for carpeta in obtener_carpetas_a_organizar(reglas):
        if not carpeta.is_dir():
            logging.error(f"No se encontró la carpeta: {carpeta}")
            continue
        logging.info(f"Organizando: {carpeta}")
        organizar_carpeta(carpeta, reglas)

    logging.info("Listo.")


if __name__ == "__main__":
    main()
